"""Tests for file writer module."""
import tempfile
from datetime import datetime
from pathlib import Path
from unittest.mock import patch

import pandas as pd
import pytest
from openpyxl import load_workbook

from src.file_writer import FileWriter


class TestFileWriter:
    """Test file writer functionality."""
    
    @pytest.fixture
    def temp_output_dir(self):
        """Create temporary output directories."""
        with tempfile.TemporaryDirectory() as tmpdir:
            base_path = Path(tmpdir)
            dailies_dir = base_path / "dailies"
            tables_dir = base_path / "tables"
            dailies_dir.mkdir(parents=True)
            tables_dir.mkdir(parents=True)
            yield base_path
    
    @pytest.fixture
    def file_writer(self, temp_output_dir):
        """Create file writer with temporary directories."""
        with patch('src.file_writer.settings') as mock_settings:
            mock_settings.outputs_dir = temp_output_dir
            mock_settings.dailies_dir = temp_output_dir / "dailies"
            mock_settings.tables_dir = temp_output_dir / "tables"
            mock_settings.ensure_directories = lambda: None
            return FileWriter()
    
    def test_get_date_string(self, file_writer):
        """Test date string formatting."""
        # Test with specific date
        test_date = datetime(2025, 7, 9)
        date_str = file_writer._get_date_string(test_date)
        assert date_str == "07.09.25"
        
        # Test with None (should use today)
        date_str = file_writer._get_date_string(None)
        assert len(date_str) == 8
        assert date_str.count('.') == 2
    
    def test_write_daily_txt(self, file_writer, temp_output_dir):
        """Test writing daily TXT file."""
        content = "This is the daily briefing content.\nLine 2\nLine 3"
        test_date = datetime(2025, 7, 9)
        
        filepath = file_writer.write_daily_txt(content, test_date)
        
        # Check file was created
        assert filepath.exists()
        assert filepath.name == "Daily_07.09.25.txt"
        assert filepath.parent == temp_output_dir / "dailies"
        
        # Check content
        written_content = filepath.read_text(encoding='utf-8')
        assert written_content == content
    
    def test_create_initial_xlsx(self, file_writer):
        """Test creating initial XLSX file."""
        headers = ["Date", "Field1", "Field2", "Field3"]
        
        wb = file_writer._create_initial_xlsx(headers)
        ws = wb.active
        
        # Check headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            assert cell.value == header
            assert cell.font.bold  # Headers should be bold
        
        # Check worksheet title
        assert ws.title == "Daily Briefings"
    
    def test_update_table_xlsx_new_file(self, file_writer, temp_output_dir):
        """Test creating new XLSX table."""
        context = {
            "FIELD1": "Value 1",
            "FIELD2": "Value 2",
            "FIELD3": "Value 3",
        }
        test_date = datetime(2025, 7, 9)
        
        filepath = file_writer.update_table_xlsx(context, test_date)
        
        # Check file was created
        assert filepath.exists()
        assert filepath.name == "Table_07.09.25.xlsx"
        
        # Load and check content
        wb = load_workbook(filepath)
        ws = wb.active
        
        # Check headers (row 1)
        assert ws.cell(row=1, column=1).value == "Date"
        assert ws.cell(row=1, column=2).value == "FIELD1"
        
        # Check data (row 2)
        assert ws.cell(row=2, column=1).value == "2025-07-09"
        assert ws.cell(row=2, column=2).value == "Value 1"
        assert ws.cell(row=2, column=3).value == "Value 2"
    
    def test_update_table_xlsx_append(self, file_writer, temp_output_dir):
        """Test appending to existing XLSX table."""
        # Create initial file
        context1 = {
            "FIELD1": "Day 1 Value",
            "FIELD2": "Day 1 Value 2",
        }
        date1 = datetime(2025, 7, 8)
        filepath1 = file_writer.update_table_xlsx(context1, date1)
        
        # Update with new day's data
        context2 = {
            "FIELD1": "Day 2 Value",
            "FIELD2": "Day 2 Value 2",
        }
        date2 = datetime(2025, 7, 9)
        filepath2 = file_writer.update_table_xlsx(context2, date2)
        
        # Load and check
        wb = load_workbook(filepath2)
        ws = wb.active
        
        # Should have 3 rows (header + 2 data rows)
        assert ws.max_row == 3
        
        # Check both days' data
        assert ws.cell(row=2, column=1).value == "2025-07-08"
        assert ws.cell(row=2, column=2).value == "Day 1 Value"
        assert ws.cell(row=3, column=1).value == "2025-07-09"
        assert ws.cell(row=3, column=2).value == "Day 2 Value"
    
    def test_get_headers_from_context(self, file_writer):
        """Test extracting headers from context."""
        context = {
            "CUSTOM_FIELD": "value",
            "FIELD1": "value1",
            "ANOTHER_FIELD": "value2",
        }
        
        headers = file_writer._get_headers_from_context(context)
        
        # Should start with Date
        assert headers[0] == "Date"
        
        # Should include all context keys
        assert "CUSTOM_FIELD" in headers
        assert "FIELD1" in headers
        assert "ANOTHER_FIELD" in headers
    
    def test_prepare_row_data(self, file_writer):
        """Test preparing row data for XLSX."""
        context = {
            "FIELD1": "Text value",
            "FIELD2": 123,  # Number
            "FIELD3": None,  # None value
            "FIELD4": ["list", "value"],  # Non-string
        }
        test_date = datetime(2025, 7, 9)
        
        row_data = file_writer._prepare_row_data(context, test_date)
        
        # First value should be date
        assert row_data[0] == "2025-07-09"
        
        # All values should be strings
        assert all(isinstance(val, str) for val in row_data)
        
        # Check conversions
        assert "Text value" in row_data
        assert "123" in row_data
        assert "['list', 'value']" in row_data
    
    def test_get_missing_fields(self, file_writer):
        """Test identifying missing fields."""
        context = {
            "FIELD1": "Has value",
            "FIELD2": "",  # Empty string
            "FIELD3": "(data unavailable)",  # Standard missing marker
            "FIELD4": "Another value",
            "FIELD5": None,  # None
        }
        
        missing = file_writer.get_missing_fields(context)
        
        assert "FIELD2" in missing  # Empty string
        assert "FIELD3" in missing  # Standard marker
        assert "FIELD5" in missing  # None
        assert "FIELD1" not in missing  # Has value
        assert "FIELD4" not in missing  # Has value
        assert len(missing) == 3
    
    def test_write_daily_txt_error_handling(self, file_writer):
        """Test error handling in write_daily_txt."""
        # Try to write to invalid path
        with patch('src.file_writer.settings') as mock_settings:
            mock_settings.dailies_dir = Path("/invalid/path/that/does/not/exist")
            file_writer = FileWriter()
            
            with pytest.raises(Exception):
                file_writer.write_daily_txt("content")