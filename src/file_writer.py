"""File writer module for saving daily briefings and Excel tables."""
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .settings import settings
from .utils.logger import get_logger

logger = get_logger(__name__)


class FileWriter:
    """Handles writing daily briefings to TXT and XLSX files."""
    
    def __init__(self):
        """Initialize the file writer."""
        # Ensure output directories exist
        settings.ensure_directories()
        
    def _get_date_string(self, date: Optional[datetime] = None) -> str:
        """Get formatted date string for filenames.
        
        Args:
            date: Date to format (defaults to today)
            
        Returns:
            Formatted date string (MM.DD.YY)
        """
        if date is None:
            date = datetime.now()
        return date.strftime("%m.%d.%y")
    
    def write_daily_txt(self, content: str, date: Optional[datetime] = None) -> Path:
        """Write daily briefing to TXT file.
        
        Args:
            content: Briefing content
            date: Date for the briefing (defaults to today)
            
        Returns:
            Path to the written file
        """
        date_str = self._get_date_string(date)
        filename = f"Daily_{date_str}.txt"
        filepath = settings.dailies_dir / filename
        
        try:
            filepath.write_text(content, encoding='utf-8')
            logger.info(
                "Daily TXT file written",
                filepath=str(filepath),
                size=len(content),
            )
            return filepath
            
        except Exception as e:
            logger.error(
                "Failed to write daily TXT file",
                filepath=str(filepath),
                error=str(e),
            )
            raise
    
    def _get_latest_xlsx_file(self) -> Optional[Path]:
        """Get the most recent XLSX file from the tables directory.
        
        Returns:
            Path to the latest XLSX file, or None if no files exist
        """
        xlsx_files = list(settings.tables_dir.glob("Table_*.xlsx"))
        if not xlsx_files:
            return None
            
        # Sort by modification time to get the most recent
        return max(xlsx_files, key=lambda p: p.stat().st_mtime)
    
    def _create_initial_xlsx(self, headers: List[str]) -> Workbook:
        """Create initial XLSX file with headers.
        
        Args:
            headers: List of column headers
            
        Returns:
            New workbook with formatted headers
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Daily Briefings"
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            # Format header cell
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
            
        # Date column should be wider
        ws.column_dimensions['A'].width = 12
        
        logger.info("Created initial XLSX file with headers", headers_count=len(headers))
        return wb
    
    def update_table_xlsx(self, context: Dict[str, Any], date: Optional[datetime] = None) -> Path:
        """Update the cumulative XLSX table with new daily data.
        
        Args:
            context: Dictionary of all briefing data
            date: Date for the briefing (defaults to today)
            
        Returns:
            Path to the updated XLSX file
        """
        date_obj = date or datetime.now()
        date_str = self._get_date_string(date_obj)
        filename = f"Table_{date_str}.xlsx"
        filepath = settings.tables_dir / filename
        
        try:
            # Get the latest XLSX file
            latest_file = self._get_latest_xlsx_file()
            
            if latest_file and latest_file.name != filename:
                # Copy the latest file to today's file
                shutil.copy2(latest_file, filepath)
                wb = load_workbook(filepath)
                ws = wb.active
                logger.info(
                    "Copied existing XLSX file",
                    source=str(latest_file),
                    destination=str(filepath),
                )
            else:
                # Create new file with headers
                headers = self._get_headers_from_context(context)
                wb = self._create_initial_xlsx(headers)
                ws = wb.active
            
            # Add new row with today's data
            row_data = self._prepare_row_data(context, date_obj)
            ws.append(row_data)
            
            # Save the workbook
            wb.save(filepath)
            logger.info(
                "XLSX table updated",
                filepath=str(filepath),
                row_number=ws.max_row,
            )
            return filepath
            
        except Exception as e:
            logger.error(
                "Failed to update XLSX table",
                filepath=str(filepath),
                error=str(e),
            )
            raise
    
    def _get_headers_from_context(self, context: Dict[str, Any]) -> List[str]:
        """Extract headers from context in template order.
        
        Args:
            context: Dictionary of briefing data
            
        Returns:
            List of headers starting with Date
        """
        # Define the expected order based on the template
        ordered_fields = [
            "Date",
            "FULLDATE",
            "YC_ARTICLE_PICK",
            "YC_ARTICLE_SUMMARY",
            "YC_ARTICLE_KEYPOINTS",
            "YC_ARTICLE_KEYWORDS",
            "GITHUB_TRENDING_MCP_NAME",
            "GITHUB_TRENDING_MCP_SUMMARY",
            "TRANSCRIPT_TABLE",
            "COUNTRY_OF_THE_DAY",
            "COUNTRY_CAPITAL_OF_THE_DAY",
            "CAPITAL_LOCATION_BREAKDOWN",
            "GET_TO_IT_SAYING",
            "CODEBASE_TODAY",
            "CODEBASE_SUMMARY",
            "DAYS_LEFT",
            "CURRENT_STUDY_YEAR",
            "CURRENT_STUDY_STATE",
            "CURRENT_YEAR_BEST_PICTURE",
            "CURRENT_YEAR_BEST_ACTOR_IN_PICTURE",
            "CURRENT_YEAR_BEST_CINEMATOGRAPHY",
            "CURRENT_YEAR_BEST_SCORE",
            "CURRENT_YEAR_BEST_FOREIGN_FILM",
            "MAJOR_INVENTION_OF_YEAR",
            "MAJOR_INVENTION_SUMMARY",
            "CURRENT_GOLF_STATE_SUMMARY",
            "CURRENT_YEAR_US_PRESIDENT_VPS",
            "NEW_MAJOR_PRESIDENTIAL_DECISION",
            "WW1_FACT",
            "WW2_FACT",
            "EUROPE_FACT",
            "IRELAND_FACT",
            "JERUSALEM_FACT",
            "INDIA_FACT",
            "MEXICO_FACT",
            "STUNT_RIGGING_SUMMARY",
            "BIKE_FUN_FACT",
            "NASA_LAUNCH_HISTORY",
            "TEACH_ME_GC",
            "QUIZ_ME_CS_TERM",
            "QUIZ_ME_ESPANOL",
        ]
        
        # Add any additional fields from context that aren't in our list
        for key in context.keys():
            if key not in ordered_fields:
                ordered_fields.append(key)
                
        return ordered_fields
    
    def _prepare_row_data(self, context: Dict[str, Any], date: datetime) -> List[Any]:
        """Prepare row data for XLSX in the correct order.
        
        Args:
            context: Dictionary of briefing data
            date: Date for the briefing
            
        Returns:
            List of values in header order
        """
        headers = self._get_headers_from_context(context)
        row_data = []
        
        for header in headers:
            if header == "Date":
                # Special handling for date column
                row_data.append(date.strftime("%Y-%m-%d"))
            else:
                # Get value from context, use empty string if missing
                value = context.get(header, "")
                # Convert non-string values to string
                if not isinstance(value, str):
                    value = str(value)
                row_data.append(value)
                
        return row_data
    
    def get_missing_fields(self, context: Dict[str, Any]) -> List[str]:
        """Get list of fields that have missing or empty data.
        
        Args:
            context: Dictionary of briefing data
            
        Returns:
            List of field names with missing data
        """
        missing_fields = []
        
        for key, value in context.items():
            if not value or value == "(data unavailable)":
                missing_fields.append(key)
                
        return missing_fields